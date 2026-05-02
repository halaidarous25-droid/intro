"""
نظام توزيع المسؤوليات الذكي — Flask Backend
"""
import json, uuid, io
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
from anthropic import Anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

app = Flask(__name__)
client = Anthropic()

DATA_FILE = Path(__file__).parent / "data.json"

def load_data():
    if DATA_FILE.exists():
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    return {"employees": [], "responsibilities": []}

def save_data(data):
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def new_id(): return str(uuid.uuid4())[:8]

COLORS = ["#0f766e","#7c3aed","#0369a1","#b45309","#be185d","#15803d","#b91c1c","#0e7490"]

def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg="0F766E"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin()

def parse_json_safe(text):
    import re
    try: return json.loads(text)
    except Exception: pass
    for pat in [r'```json\s*([\s\S]*?)```', r'```\s*([\s\S]*?)```']:
        m = re.search(pat, text, re.I)
        if m:
            try: return json.loads(m.group(1).strip())
            except Exception: pass
    depth=0; start=-1
    for i,ch in enumerate(text):
        if ch=='{':
            if start==-1: start=i
            depth+=1
        elif ch=='}' and depth>0:
            depth-=1
            if depth==0:
                try: return json.loads(text[start:i+1])
                except Exception: start=-1
    return None

# ── Pages ─────────────────────────────────────────────────────

@app.route("/")
def index(): return render_template("index.html")

# ── Employees ─────────────────────────────────────────────────

@app.route("/api/employees", methods=["GET"])
def get_employees(): return jsonify(load_data()["employees"])

@app.route("/api/employees", methods=["POST"])
def create_employee():
    data = load_data()
    emp = {**request.json, "id": new_id(),
           "color": COLORS[len(data["employees"]) % len(COLORS)],
           "aiAnalysis": request.json.get("aiAnalysis")}
    data["employees"].append(emp)
    save_data(data); return jsonify(emp), 201

@app.route("/api/employees/<eid>", methods=["PUT"])
def update_employee(eid):
    data = load_data()
    for i,e in enumerate(data["employees"]):
        if e["id"]==eid:
            data["employees"][i]={**e,**request.json,"id":eid}
            save_data(data); return jsonify(data["employees"][i])
    return jsonify({"error":"not found"}),404

@app.route("/api/employees/<eid>", methods=["DELETE"])
def delete_employee(eid):
    data=load_data(); data["employees"]=[e for e in data["employees"] if e["id"]!=eid]
    save_data(data); return jsonify({"ok":True})

# ── Responsibilities ───────────────────────────────────────────

@app.route("/api/responsibilities", methods=["GET"])
def get_responsibilities(): return jsonify(load_data()["responsibilities"])

@app.route("/api/responsibilities", methods=["POST"])
def create_responsibility():
    data=load_data()
    resp={**request.json,"id":new_id(),"assignedTo":None,"cover1":None,"cover2":None,"aiData":None}
    data["responsibilities"].append(resp); save_data(data); return jsonify(resp),201

@app.route("/api/responsibilities/<rid>", methods=["PUT"])
def update_responsibility(rid):
    data=load_data()
    for i,r in enumerate(data["responsibilities"]):
        if r["id"]==rid:
            data["responsibilities"][i]={**r,**request.json,"id":rid}
            save_data(data); return jsonify(data["responsibilities"][i])
    return jsonify({"error":"not found"}),404

@app.route("/api/responsibilities/<rid>", methods=["DELETE"])
def delete_responsibility(rid):
    data=load_data(); data["responsibilities"]=[r for r in data["responsibilities"] if r["id"]!=rid]
    save_data(data); return jsonify({"ok":True})

# ── AI: Analyze Employee ──────────────────────────────────────

@app.route("/api/ai/analyze-employee", methods=["POST"])
def analyze_employee():
    body=request.json; bio=body.get("bio",""); skills=body.get("skills","")
    resp=client.messages.create(
        model="claude-sonnet-4-5", max_tokens=1000,
        system="You are an HR expert. Return ONLY valid JSON.",
        messages=[{"role":"user","content":
            f"Analyze employee:\nDescription: {bio or 'none'}\nSkills: {skills or 'none'}\n\n"
            'Return ONLY JSON: {"personality_type":"…","strengths":["…","…","…","…"],'
            '"weaknesses":["…","…","…"],"key_skills":["…","…","…","…","…"],'
            '"soft_skills":["…","…","…"],"work_style":"…","best_for":["…","…","…"],'
            '"avoid":["…","…"],"summary":"…","confidence":80}'}])
    result=parse_json_safe(resp.content[0].text)
    return jsonify(result or {"error":"parse failed"})

# ── AI: Analyze Responsibility ────────────────────────────────

@app.route("/api/ai/analyze-responsibility", methods=["POST"])
def analyze_responsibility():
    body=request.json; name=body.get("name",""); desc=body.get("desc","")
    resp=client.messages.create(
        model="claude-sonnet-4-5", max_tokens=900,
        system="You are an HR expert. Return ONLY valid JSON.",
        messages=[{"role":"user","content":
            f"Analyze job responsibility:\nName: {name or 'unspecified'}\nDesc: {desc or 'none'}\n\n"
            'Return ONLY JSON: {"weight":3,"load":3,"impact":"medium","complexity":"moderate",'
            '"freq":10,"time_minutes":30,"category":"other","systems":["…"],'
            '"req_skills":["…","…"],"req_traits":["…","…"],'
            '"sub_tasks":["…","…","…"],"kpi":"…","risk":"…","summary":"…"}'}])
    result=parse_json_safe(resp.content[0].text)
    return jsonify(result or {"error":"parse failed"})

# ── AI: Distribute ────────────────────────────────────────────

@app.route("/api/ai/distribute", methods=["POST"])
def distribute():
    data=load_data(); emps=data["employees"]; resps=data["responsibilities"]
    if not emps or not resps:
        return jsonify({"error":"يحتاج عضو واحد ومسؤولية واحدة على الأقل"}),400

    emp_lines="\n".join(
        f"ID:{e['id']} | {e['name']} | {e.get('title','')} | {e.get('exp',0)}yrs | "
        f"Skills:{','.join((e.get('aiAnalysis') or {}).get('key_skills',[]))} | "
        f"Bio:{((e.get('aiAnalysis') or {}).get('summary','') or e.get('bio',''))[:100]}"
        for e in emps)
    resp_lines="\n".join(
        f"ID:{r['id']} | {r['name']} | W:{r.get('weight',3)} L:{r.get('load',3)} | "
        f"Impact:{r.get('impact','medium')} | Skills:{','.join(r.get('reqSkills',[]))} | "
        f"Desc:{r.get('desc','')[:120]}"
        for r in resps)

    prompt=(f"TEAM:\n{emp_lines}\n\nRESPONSIBILITIES:\n{resp_lines}\n\n"
            "Return ONLY JSON with real IDs:\n"
            '{"analysis":[{"resp_id":"ID","owner_id":"ID","cover1_id":"ID","cover2_id":"ID",'
            '"rankings":[{"emp_id":"ID","rank":1,"score":85,"role":"Owner","reasons":["r1","r2"],"warnings":[],"fit_summary":"sentence"},'
            '{"emp_id":"ID","rank":2,"score":70,"role":"Cover1","reasons":["r"],"warnings":[],"fit_summary":"sentence"},'
            '{"emp_id":"ID","rank":3,"score":55,"role":"Cover2","reasons":["r"],"warnings":[],"fit_summary":"sentence"}],'
            '"recommendation":"advice","risk_if_wrong":"consequence"}],'
            '"team_summary":"2 sentences","load_analysis":[{"emp_id":"ID","load_pct":30,"status":"balanced","comment":"note"}],'
            '"overload_warnings":[],"suggestions":["s1","s2"]}')

    resp_ai=client.messages.create(
        model="claude-sonnet-4-5", max_tokens=4000,
        system="You are an HR specialist. Return ONLY valid JSON. Use EXACT IDs. owner/cover1/cover2 must be different.",
        messages=[{"role":"user","content":prompt}])

    result=parse_json_safe(resp_ai.content[0].text)
    if not result:
        return jsonify({"error":"Invalid response format","raw":resp_ai.content[0].text[:400]}),500

    for item in result.get("analysis",[]):
        for r in resps:
            if r["id"]==item.get("resp_id"):
                rnk=item.get("rankings",[])
                r["assignedTo"]=item.get("owner_id") or (rnk[0].get("emp_id") if rnk else None)
                r["cover1"]=item.get("cover1_id") or (rnk[1].get("emp_id") if len(rnk)>1 else None)
                r["cover2"]=item.get("cover2_id") or (rnk[2].get("emp_id") if len(rnk)>2 else None)
    data["responsibilities"]=resps; save_data(data)
    return jsonify(result)

# ── Import: Employees ─────────────────────────────────────────

@app.route("/api/import/employees", methods=["POST"])
def import_employees():
    file=request.files.get("file")
    if not file: return jsonify({"error":"no file"}),400
    wb=openpyxl.load_workbook(file,data_only=True)
    ws=next((wb[s] for s in wb.sheetnames if "دليل" not in s),wb.active)
    rows=list(ws.iter_rows(values_only=True))
    hi=next((i for i,row in enumerate(rows[:8]) if sum(1 for c in row if c)>=3),0)
    headers=[str(c or "").strip().replace("*","").strip() for c in rows[hi]]
    data=load_data(); added=0

    def gv(row,*keys):
        for k in keys:
            for ci,h in enumerate(headers):
                if k in h and ci<len(row) and row[ci]:
                    return str(row[ci]).strip()
        return ""

    for row in rows[hi+1:]:
        if not any(row): continue
        name=gv(row,"الاسم")
        if not name or len(name)<2: continue
        emp={"id":new_id(),"name":name,"title":gv(row,"المسمى","الوظيفي"),
             "dept":gv(row,"القسم","الفريق"),"exp":int(gv(row,"الخبرة","سنوات") or 0),
             "bio":gv(row,"وصف","شخصية"),"skills":gv(row,"مهارات","خبرات"),
             "color":COLORS[len(data["employees"])%len(COLORS)],"aiAnalysis":None}
        data["employees"].append(emp); added+=1
    save_data(data)
    return jsonify({"added":added,"employees":data["employees"]})

# ── Import: Responsibilities ───────────────────────────────────

@app.route("/api/import/responsibilities", methods=["POST"])
def import_responsibilities():
    file=request.files.get("file")
    if not file: return jsonify({"error":"no file"}),400
    wb=openpyxl.load_workbook(file,data_only=True)
    ws=next((wb[s] for s in wb.sheetnames if "دليل" not in s),wb.active)
    rows=list(ws.iter_rows(values_only=True))
    hi=next((i for i,row in enumerate(rows[:8]) if sum(1 for c in row if c)>=3),0)
    headers=[str(c or "").strip().replace("*","").strip() for c in rows[hi]]
    data=load_data(); added=0
    CMAP={"استقطاب":"onboarding","إنهاء":"offboarding","إجازات":"leave","رواتب":"payroll",
          "امتثال":"compliance","تدريب":"training","أداء":"performance"}
    IMAP={"منخفض":"low","متوسط":"medium","عالٍ":"high","حرج":"critical"}

    def gv(row,*keys):
        for k in keys:
            for ci,h in enumerate(headers):
                if k in h and ci<len(row) and row[ci]:
                    return str(row[ci]).strip()
        return ""

    for row in rows[hi+1:]:
        if not any(row): continue
        name=gv(row,"اسم المسؤولية","الاسم")
        if not name or len(name)<2: continue
        cat_r=gv(row,"الفئة"); imp_r=gv(row,"التأثير","مستوى")
        sys_s=gv(row,"الأنظمة"); sk_s=gv(row,"المهارات")
        resp={"id":new_id(),"name":name,"desc":gv(row,"الوصف"),
              "category":next((v for k,v in CMAP.items() if k in cat_r),"other"),
              "impact":next((v for k,v in IMAP.items() if k in imp_r),"medium"),
              "weight":int(gv(row,"الوزن") or 3),"load":int(gv(row,"العبء") or 3),
              "freq":int(gv(row,"التكرار") or 0) or None,
              "time":int(gv(row,"الوقت") or 0) or None,
              "systems":[s.strip() for s in sys_s.split(",") if s.strip()] if sys_s else [],
              "reqSkills":[s.strip() for s in sk_s.split(",") if s.strip()] if sk_s else [],
              "reqTraits":[],"assignedTo":None,"cover1":None,"cover2":None,"aiData":None}
        data["responsibilities"].append(resp); added+=1
    save_data(data)
    return jsonify({"added":added,"responsibilities":data["responsibilities"]})

# ── Templates Download ─────────────────────────────────────────

@app.route("/api/template/employees")
def template_employees():
    wb=openpyxl.Workbook(); ws=wb.active
    ws.title="بيانات الموظفين"; ws.sheet_view.rightToLeft=True
    ws.merge_cells("A1:G1")
    ws["A1"].value="نموذج استيراد بيانات أعضاء الفريق"
    ws["A1"].font=Font(name="Arial",bold=True,size=14,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor="0F766E")
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=40
    ws.merge_cells("A2:G2")
    ws["A2"].value="الحقول ذات الخلفية البرتقالية مطلوبة — احذف صفوف الأمثلة"
    ws["A2"].font=Font(name="Arial",size=9,italic=True,color="92400E")
    ws["A2"].fill=PatternFill("solid",fgColor="FEF3C7")
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
    cols=[("الاسم الكامل *",24,True),("المسمى الوظيفي *",24,True),("القسم / الفريق",22,False),
          ("سنوات الخبرة",13,False),("وصف الشخصية والسلوك",50,False),
          ("الخبرات والمهارات التقنية",50,False),("ملاحظات",25,False)]
    for i,(name,w,req) in enumerate(cols,1):
        hdr(ws,3,i,name,"D97706" if req else "0F766E")
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=32
    examples=[("أحمد العمري","أخصائي موارد بشرية","الاستحقاقات",5,"دقيق، ممتاز تحت الضغط","Oracle HCM، GOSI، Excel",""),
              ("سارة الزهراني","مشرفة إجازات","دورة الموظف",8,"منهجية، دقيقة في الإجراءات","Oracle HCM، ECTS","")]
    for ri,ex in enumerate(examples,4):
        for ci,val in enumerate(ex,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.font=Font(name="Arial",size=9,italic=True,color="7C8B9A")
            c.fill=PatternFill("solid",fgColor="F8FAFC")
            c.alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
            c.border=thin()
    for row in range(6,52):
        for col in range(1,8):
            c=ws.cell(row=row,column=col,value="")
            c.fill=PatternFill("solid",fgColor="FFF7ED" if col<=2 else ("FFFFFF" if row%2==0 else "F8FAFC"))
            c.alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
            c.font=Font(name="Arial",size=10); c.border=thin()
        ws.row_dimensions[row].height=30
    ws.freeze_panes="A4"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name="قالب_بيانات_الموظفين.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/template/responsibilities")
def template_responsibilities():
    wb=openpyxl.Workbook(); ws=wb.active
    ws.title="المسؤوليات"; ws.sheet_view.rightToLeft=True
    ws.merge_cells("A1:L1")
    ws["A1"].value="نموذج استيراد المسؤوليات الوظيفية"
    ws["A1"].font=Font(name="Arial",bold=True,size=14,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor="0F766E")
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=40
    ws.merge_cells("A2:L2")
    ws["A2"].value="أدخل الاسم والوصف فقط — يستخرج AI باقي الحقول"
    ws["A2"].font=Font(name="Arial",size=9,italic=True,color="92400E")
    ws["A2"].fill=PatternFill("solid",fgColor="FEF3C7")
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
    cols=[("اسم المسؤولية *",30,True),("الوصف التفصيلي",55,False),
          ("الفئة",20,False),("مستوى التأثير",15,False),("الوزن 1-5",11,False),
          ("العبء 1-5",11,False),("التكرار/شهر",13,False),("الوقت/مرة (دقيقة)",15,False),
          ("الأنظمة (بفاصلة)",35,False),("المهارات (بفاصلة)",35,False),
          ("سمات الشخصية",25,False),("ملاحظات",22,False)]
    for i,(name,w,req) in enumerate(cols,1):
        hdr(ws,3,i,name,"D97706" if req else "0F766E")
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=38
    dv=DataValidation(type="list",formula1='"استقطاب وتعيين,إنهاء خدمة,إجازات وغياب,رواتب واستحقاقات,امتثال وتقارير,تدريب وتطوير,أداء وتقييم,أخرى"',allow_blank=True)
    dv.sqref="C4:C200"; ws.add_data_validation(dv)
    dv2=DataValidation(type="list",formula1='"منخفض,متوسط,عالٍ,حرج"',allow_blank=True)
    dv2.sqref="D4:D200"; ws.add_data_validation(dv2)
    examples=[("معالجة طلبات الإجازة","40 طلب شهرياً/15 دقيقة. استقبال، تحقق من رصيد، تحديث Oracle","إجازات وغياب","متوسط",3,2,40,15,"Oracle HCM","دقة, Oracle","منظم",""),
              ("احتساب الرواتب","مرة شهرياً/3 أيام. ECTS+GOSI+المالية","رواتب واستحقاقات","حرج",5,5,1,1440,"Oracle,ECTS,GOSI","محاسبة","دقيق","")]
    for ri,ex in enumerate(examples,4):
        for ci,val in enumerate(ex,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.font=Font(name="Arial",size=9,italic=True,color="7C8B9A")
            c.fill=PatternFill("solid",fgColor="F8FAFC")
            c.alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
            c.border=thin()
    for row in range(6,52):
        for col in range(1,13):
            c=ws.cell(row=row,column=col,value="")
            c.fill=PatternFill("solid",fgColor="FFF7ED" if col==1 else ("FFFFFF" if row%2==0 else "F8FAFC"))
            c.alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
            c.font=Font(name="Arial",size=10); c.border=thin()
        ws.row_dimensions[row].height=32
    ws.freeze_panes="A4"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name="قالب_بيانات_المسؤوليات.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Export Report ─────────────────────────────────────────────

@app.route("/api/export/report")
def export_report():
    data=load_data(); emps={e["id"]:e for e in data["employees"]}
    resps=data["responsibilities"]; total=len(resps) or 1
    IMP_AR={"low":"منخفض","medium":"متوسط","high":"عالٍ","critical":"حرج"}
    CATS_AR={"onboarding":"استقطاب","offboarding":"إنهاء خدمة","leave":"إجازات",
             "payroll":"رواتب","compliance":"امتثال","training":"تدريب","performance":"أداء","other":"أخرى"}
    IMP_CLR={"low":"DCFCE7","medium":"DBEAFE","high":"FEF3C7","critical":"FEE2E2"}
    wb=openpyxl.Workbook()
    # Sheet1: Distribution
    ws1=wb.active; ws1.title="ملخص التوزيع"; ws1.sheet_view.rightToLeft=True
    ws1.merge_cells("A1:I1")
    ws1["A1"].value="تقرير توزيع المسؤوليات"
    ws1["A1"].font=Font(name="Arial",bold=True,size=14,color="FFFFFF")
    ws1["A1"].fill=PatternFill("solid",fgColor="0F766E")
    ws1["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height=38
    h1=["المسؤولية","الفئة","التأثير","الوزن","العبء","صاحب المسؤولية","المغطي الرئيسي","المغطي الاحتياطي","ساعات/شهر"]
    for i,h in enumerate(h1,1):
        hdr(ws1,2,i,h)
        ws1.column_dimensions[get_column_letter(i)].width=[30,18,12,9,9,22,22,22,12][i-1]
    for ri,r in enumerate(resps,3):
        hrs=round((r.get("freq") or 0)*(r.get("time") or 0)/60,1) if r.get("freq") and r.get("time") else "—"
        row_data=[r.get("name",""),CATS_AR.get(r.get("category",""),"أخرى"),
                  IMP_AR.get(r.get("impact","medium"),"متوسط"),r.get("weight",3),r.get("load",3),
                  emps.get(r.get("assignedTo") or "","").get("name","—") if isinstance(emps.get(r.get("assignedTo") or ""),dict) else "—",
                  emps.get(r.get("cover1") or "","").get("name","—") if isinstance(emps.get(r.get("cover1") or ""),dict) else "—",
                  emps.get(r.get("cover2") or "","").get("name","—") if isinstance(emps.get(r.get("cover2") or ""),dict) else "—",
                  hrs]
        bg=IMP_CLR.get(r.get("impact","medium"),"FFFFFF")
        for ci,val in enumerate(row_data,1):
            c=ws1.cell(row=ri,column=ci,value=val)
            c.font=Font(name="Arial",size=10)
            c.fill=PatternFill("solid",fgColor=bg if ci<=5 else "FFFFFF")
            c.alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
            c.border=thin()
        ws1.row_dimensions[ri].height=26
    # Sheet2: Team Load
    ws2=wb.create_sheet("أعباء الفريق"); ws2.sheet_view.rightToLeft=True
    ws2.merge_cells("A1:G1")
    ws2["A1"].value="توزيع الأعباء على الفريق"
    ws2["A1"].font=Font(name="Arial",bold=True,size=14,color="FFFFFF")
    ws2["A1"].fill=PatternFill("solid",fgColor="0F766E")
    ws2["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=38
    h2=["الموظف","المسمى الوظيفي","مسؤولياته","% من الإجمالي","مغطي رئيسي","مغطي احتياطي","الحالة"]
    for i,h in enumerate(h2,1):
        hdr(ws2,2,i,h)
        ws2.column_dimensions[get_column_letter(i)].width=[28,25,14,14,14,14,14][i-1]
    for ri,emp in enumerate(data["employees"],3):
        owned=sum(1 for r in resps if r.get("assignedTo")==emp["id"])
        c1=sum(1 for r in resps if r.get("cover1")==emp["id"])
        c2=sum(1 for r in resps if r.get("cover2")==emp["id"])
        pct=round(owned/total*100,1)
        status="محمّل" if pct>60 else ("خفيف" if pct<15 else "متوازن")
        st_clr={"محمّل":"FEE2E2","متوازن":"DCFCE7","خفيف":"DBEAFE"}[status]
        for ci,val in enumerate([emp.get("name",""),emp.get("title",""),owned,f"{pct}%",c1,c2,status],1):
            c=ws2.cell(row=ri,column=ci,value=val)
            c.font=Font(name="Arial",size=10,bold=(ci==1))
            c.fill=PatternFill("solid",fgColor=st_clr if ci==7 else "FFFFFF")
            c.alignment=Alignment(horizontal="center" if ci>2 else "right",vertical="center")
            c.border=thin()
        ws2.row_dimensions[ri].height=26
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name="تقرير_توزيع_المسؤوليات.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    app.run(debug=True,port=5000)
